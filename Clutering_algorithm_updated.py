import pandas as pd
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans
import os
import matplotlib.pyplot as plt
from kneed import KneeLocator
from sklearn.metrics import silhouette_score
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import xml.etree.ElementTree as ET
from xml.dom import minidom
import datetime
import shutil
import pickle
import csv
import confluence_get_page
path=r'\\rover\cts\Axiom\Executables\ModuleVsTechArea_Data'
# Data
output_file_path = os.path.join(os.getcwd(), 'categorized_data.xlsx')
file_path = os.path.join(os.getcwd(), 'Final_TestCases_Data.xlsx')
print(file_path)
excel_data = pd.read_excel(file_path, sheet_name=None)
def check_for_file(file):
    if os.path.isdir(file):
        # Remove the directory and all its contents
        shutil.rmtree(file)
    elif os.path.exists(os.path.join(os.getcwd(), file)):
        print(f"Removing : {os.path.join(os.getcwd(), file)}")
        os.remove(os.path.join(os.getcwd(), file))

def parse_confluence_data():
    reference_data=[]
    for root , dirs , files in os.walk(path):
        for file in files:
            if file.endswith(".csv"):
                with open(os.path.join(path,file),'r',encoding='utf-8') as csv_file:
                    csvreader=csv.reader(csv_file)
                    for rows in csvreader:
                        reference_data.append([rows[0],rows[4]])
    return reference_data
def create_subplan(filtered_df, suite, clusters):
    cluster_wise_count=0
    module_list = filtered_df['Module'].tolist()
    if len(module_list) != 0:
        for modules in module_list:
            for rows in parse_confluence_data():
                # print(rows)
                if modules == rows[0]:
                    cluster_wise_count += int(rows[1])
        new_root = minidom.Document()
        xml = new_root.createElement("Subplan")
        xml.setAttribute("version", "2.0")
        new_root.appendChild(xml)
        for module in module_list:
            entry = new_root.createElement("Entry")
            entry.setAttribute("include", f"{module}")
            xml.appendChild(entry)
        xml_str = new_root.toprettyxml(indent="\t")
        return xml_str,cluster_wise_count
    else:
        return 0,0
def convert_ms_to_hms(milliseconds):
    seconds = milliseconds // 1000
    milliseconds = milliseconds % 1000
    minutes = seconds // 60
    seconds = seconds % 60
    hours = minutes // 60
    minutes = minutes % 60

    return hours, minutes, seconds, milliseconds
def calculate_time_taken(filtered_df,suite,clusters):
    time_taken=0
    sample_time=0
    global loaded_data
    module_list=filtered_df['Module'].tolist()
    for i in range(len(module_list)):
        module_list[i]+=f">>{suite}"
    if os.path.exists(os.path.join(os.getcwd(),"time_consumption")):
        with open("time_consumption",'rb') as f:
             loaded_data=pickle.load(f)
        f.close()
        loaded_data_list=list(loaded_data.keys())
        for i in range(len(module_list)):
            for j in range(len(loaded_data_list)):
                if (module_list[i] == loaded_data_list[j]) and (suite == loaded_data_list[j].split(">>")[1]):
                    time_taken+=loaded_data[loaded_data_list[j]]
    else:
        raise Exception("Time Consumption dataset not found for further processing !!!!")
    return [suite, f"Cluster-{clusters}", time_taken,f"{convert_ms_to_hms(time_taken)[0]}hrs {convert_ms_to_hms(time_taken)[1]}mins {convert_ms_to_hms(time_taken)[2]}secs {convert_ms_to_hms(time_taken)[3]}ms"]

check_for_file("categorized_data.xlsx")
# Create DataFrame

with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    for sheet_name, df in excel_data.items():
        if sheet_name == "Statistics":
            # Select features for clustering
            # features = df[['mean', 'std_dev', 'variance', 'Coefficient of Variation']]
            features = df[['mean', 'Coefficient of Variation']]

            # Standardize the features
            scaler = StandardScaler()
            scaled_features = scaler.fit_transform(features)

            # Apply K-Means clustering

            wcss = []
            silouhette_score_list=[]
            for i in range(2, 11):
                kmeans = KMeans(n_clusters=i, init='k-means++', random_state=42)
                kmeans.fit(scaled_features)
                labels = kmeans.fit_predict(scaled_features)
                # Calculating silouhette score
                score=silhouette_score(scaled_features,labels)
                silouhette_score_list.append(score)
                wcss.append(kmeans.inertia_)

            plt.plot(range(2, 11), wcss)
            plt.title("Elbow Method")
            plt.xlabel('Number of clusters')
            plt.ylabel('WCSS')
            check_for_file("Elbow_Method.png")
            plt.savefig("Elbow_Method.png")
            plt.cla()

            plt.plot(range(2, 11), silouhette_score_list)
            plt.title("silouhette score")
            plt.xlabel('Number of clusters')
            plt.ylabel('silouhette score')
            check_for_file("silouhette.png")
            plt.savefig("silouhette.png")
            plt.cla()

            kneedle = KneeLocator(range(2, 11), wcss, curve='convex', direction='decreasing')
            optimal_clusters = kneedle.elbow
            # optimal_clusters =10
            print(f'The optimal number of clusters is: {optimal_clusters}')
            # kmeans = KMeans(n_clusters=optimal_clusters, init='k-means++', random_state=42)
            kmeans = KMeans(n_clusters=optimal_clusters, init='k-means++', random_state=42)
            df['Cluster'] = kmeans.fit_predict(scaled_features)

            cluster_centers = kmeans.cluster_centers_
            for i, (x, y) in enumerate(cluster_centers):
                print(f"Cluster {i} center: x = {x}, y = {y}")

            # segregating unique clusters
            print("*" * 20, "\n", df)
            unique_clusters = df['Cluster'].unique()
            unique_suites = df['Suite'].unique()
            now = datetime.datetime.now()
            time_stamp = now.strftime("%Y%m%d_%H%M%S")
            os.makedirs(time_stamp, exist_ok=True)
            check_for_file(f"{time_stamp}")
            time_consumption_data = []

            numeric_df = df[['mean', 'Coefficient of Variation', 'Cluster']]
            # Calculate and plot cluster centers
            centers = numeric_df.groupby('Cluster').mean()
            centers = centers.sort_values(by=['mean', 'Coefficient of Variation'], ascending=[False, True])
            prioritized_clusters = []
            for cluster, row in centers.iterrows():
                # print(f"Center of cluster {cluster}: x={row['mean']}, y={row['Coefficient of Variation']}")
                prioritized_clusters.append(cluster)
            print(f"Prioritized clusters:{prioritized_clusters}")
            for suite in unique_suites:
                suite_wise_count=0
                suite_dir = os.path.join(os.getcwd(), time_stamp, f"{suite}")
                os.makedirs(suite_dir, exist_ok=True)
                print("="*30,f"\n {suite}\n","="*30)
                for clusters in prioritized_clusters:
                    filtered_df = df[(df['Cluster'] == clusters) & (df['Suite'] == suite)]
                    my_xml_str,cluster_wise_count = create_subplan(filtered_df, suite, clusters)# Subplan creation
                    if my_xml_str != 0 and cluster_wise_count >0:
                        suite_wise_count += cluster_wise_count
                        print(f"Suite :{suite} || Cluster : {clusters}  || Total Testcases : {cluster_wise_count}")
                        with open(os.path.join(os.getcwd(), time_stamp, f"{suite}", f"P-{prioritized_clusters.index(clusters)}.xml"),'w') as f:
                            f.write(my_xml_str)
                            print(f">> P-{prioritized_clusters.index(clusters)}_{cluster_wise_count}.xml is created Successfully in {suite_dir}")

                    # Code to Calculate time for each cluster
                    time_consumption_data.append(calculate_time_taken(filtered_df, suite, clusters))
                print(f"\n *** Total testcases in {suite} is {suite_wise_count}")
                os.rename(os.path.join(os.getcwd(),time_stamp,f"{suite}"),os.path.join(os.getcwd(),time_stamp,f"{suite}_{suite_wise_count}"))
                print(f">> Renamed folder name from {suite} to  {suite}_{suite_wise_count}")

            # =========================================
            print("\n\n","="*100)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            df1 = pd.DataFrame(time_consumption_data, columns=['Suite', 'Cluster', 'Total time taken (in ms)','hr min sec ms'])
            overall_module_time=0
            for values in loaded_data.values():
                overall_module_time+=values
            hours, minutes, seconds, milliseconds = convert_ms_to_hms(overall_module_time)
            print(">"*2,f"Total time taken for all modules in xTS : {hours}hr {minutes} mins {seconds} secs {milliseconds} milli seconds")
            cluster_wise_total_time = df1['Total time taken (in ms)'].sum()
            hours, minutes, seconds, milliseconds = convert_ms_to_hms(cluster_wise_total_time)
            print(">"*2,f"Total time taken for only segregated modules : {hours}hr {minutes} mins {seconds} secs {milliseconds} milli seconds")

            time_saved=(overall_module_time-cluster_wise_total_time)
            hours, minutes, seconds, milliseconds = convert_ms_to_hms(time_saved)
            print(">"*2,f"Total time saved : {hours}hr {minutes} mins {seconds} secs {milliseconds} milli seconds")
            print("\n\n", "=" * 100,"\n")
            df1.to_excel(writer, sheet_name="Time Consumed", index=False)

            # Calculate Silhouette Score
            silhouette_avg = silhouette_score(scaled_features, df['Cluster'])
            print(f'Silhouette Score: {silhouette_avg}')
            scatter = plt.scatter(df['Coefficient of Variation'], df['mean'], c=df['Cluster'], cmap='viridis',s=75)
            plt.scatter(centers['Coefficient of Variation'], centers['mean'], c='red', marker='*', s=50)
            plt.xlabel('Consistency of Failures')
            plt.ylabel('Failure count')
            plt.title('Scatter Plot of Clusters with Centers')
            plt.legend(*scatter.legend_elements(), title="Clusters")
            plt.grid(True)
            # Remove the file if it already exists
            check_for_file('cluster_plot.png')
            plt.savefig('cluster_plot.png')
            plt.show()

        # Load the image into the Excel file
workbook = load_workbook(output_file_path)
worksheet = workbook.create_sheet("Cluster Plot")
img1 = Image('cluster_plot.png')
worksheet.add_image(img1, 'A1')
worksheet = workbook.create_sheet("Elbow Method")
img2 = Image("Elbow_Method.png")
worksheet.add_image(img2, 'A2')
workbook.save(output_file_path)

